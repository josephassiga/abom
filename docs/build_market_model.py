from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- styling helpers ----
ARIAL = "Arial"
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E2F3"
ZEBRA = "F2F5FB"

f_title = Font(name=ARIAL, size=16, bold=True, color=NAVY)
f_sub   = Font(name=ARIAL, size=10, italic=True, color="595959")
f_sect  = Font(name=ARIAL, size=12, bold=True, color=NAVY)
f_hdr   = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
f_inp   = Font(name=ARIAL, size=10, color="0000FF")   # blue inputs
f_calc  = Font(name=ARIAL, size=10, color="000000")   # black formulas
f_link  = Font(name=ARIAL, size=10, color="008000")   # green cross-sheet
f_lbl   = Font(name=ARIAL, size=10, color="000000")
f_bold  = Font(name=ARIAL, size=10, bold=True, color="000000")
f_note  = Font(name=ARIAL, size=8, italic=True, color="808080")
f_big   = Font(name=ARIAL, size=20, bold=True, color=NAVY)
f_cardlbl = Font(name=ARIAL, size=9, bold=True, color="FFFFFF")

fill_hdr  = PatternFill("solid", fgColor=BLUE)
fill_sect = PatternFill("solid", fgColor=LIGHT)
fill_zeb  = PatternFill("solid", fgColor=ZEBRA)
fill_card = PatternFill("solid", fgColor=NAVY)
fill_yellow = PatternFill("solid", fgColor="FFF2CC")

thin = Side(style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

CUR = '€#,##0;(€#,##0);"-"'
CURM = '€#,##0.0;(€#,##0.0);"-"'
CNT = '#,##0;(#,##0);"-"'
PCT = '0.0%'

def style_row_headers(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = f_hdr; cell.fill = fill_hdr; cell.border = border_all
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()

# =====================================================================
# SHEET 1: SUMMARY
# =====================================================================
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False
ws["A1"] = "ABOM — Agent Bill of Materials · bottom-up market model"; ws["A1"].font = f_title
ws["A2"] = "Agent-accountability / governance tooling for agentic AI · European regulated enterprises · all figures EUR"; ws["A2"].font = f_sub

# headline cards (filled later via links)
cards = [
    ("A4", "TAM — EU regulated agent-accountability layer (top-down)", "='Market Sizing'!B10", "A5", CURM, " €M"),
    ("D4", "SAM — bottom-up (serviceable, annual)", "='Market Sizing'!B12", "D5", CURM, " €M"),
    ("G4", "SOM — Year-5 achievable ARR (base)", "='Market Sizing'!B15", "G5", CURM, " €M"),
]
for lblcell, lbl, formula, valcell, fmt, unit in cards:
    ws[lblcell] = lbl
    ws[lblcell].font = f_cardlbl; ws[lblcell].fill = fill_card
    ws[lblcell].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws[valcell] = formula
    ws[valcell].font = f_big; ws[valcell].number_format = CURM
    ws[valcell].alignment = Alignment(horizontal="left", vertical="center")

for col in "ABCDEFGHI":
    ws.column_dimensions[col].width = 13
ws.column_dimensions["A"].width = 22
ws.column_dimensions["D"].width = 22
ws.column_dimensions["G"].width = 22
for r in (4,):
    ws.row_dimensions[r].height = 42
for r in (5,):
    ws.row_dimensions[r].height = 30

# Key outputs block
ws["A8"] = "Key outputs"; ws["A8"].font = f_sect; ws["A8"].fill = fill_sect
for c in range(1,10):
    ws.cell(row=8, column=c).fill = fill_sect
outputs = [
    ("Year-5 ABOM platform ARR (verify + Notary)", "='Revenue Build'!G12/1000", CURM, " €M"),
    ("Year-5 services revenue", "='Revenue Build'!G13/1000", CURM, " €M"),
    ("Year-5 total revenue", "='Revenue Build'!G14/1000", CURM, " €M"),
    ("Year-5 cumulative customers (logos)", "='Revenue Build'!G16", CNT, ""),
    ("Year-5 logo penetration of serviceable market", "='Revenue Build'!G17", PCT, ""),
    ("SOM as % of SAM (Year 5)", "='Market Sizing'!B16", PCT, ""),
]
r = 9
for lbl, formula, fmt, unit in outputs:
    ws.cell(row=r, column=1, value=lbl).font = f_lbl
    cell = ws.cell(row=r, column=4, value=formula)
    cell.font = f_link; cell.number_format = fmt
    r += 1

# Scenario table (logo-driven sensitivity on Year-5 ARR)
ws["A17"] = "Scenario sensitivity — Year-5 ABOM platform ARR (€M)"; ws["A17"].font = f_sect; ws["A17"].fill = fill_sect
for c in range(1,10):
    ws.cell(row=17, column=c).fill = fill_sect
ws["A18"] = "Driver: logo-ramp multiplier applied to base-case customer counts (ABOM ACV held constant)."
ws["A18"].font = f_note
style_row_headers(ws, 19, [1,2,3])
ws.cell(row=19, column=1, value="Scenario")
ws.cell(row=19, column=2, value="Logo multiplier")
ws.cell(row=19, column=3, value="Year-5 ARR (€M)")
scen = [("Conservative", 0.6), ("Base", 1.0), ("Aggressive", 1.4)]
r = 20
for name, mult in scen:
    ws.cell(row=r, column=1, value=name).font = f_lbl
    mc = ws.cell(row=r, column=2, value=mult); mc.font = f_inp; mc.number_format = '0.0"x"'
    vc = ws.cell(row=r, column=3, value=f"='Revenue Build'!$G$12/1000*B{r}")
    vc.font = f_calc; vc.number_format = CURM
    for c in (1,2,3):
        ws.cell(row=r, column=c).border = border_all
        if name == "Base":
            ws.cell(row=r, column=c).fill = fill_yellow
    r += 1

ws["A25"] = ("Note: The headline analyst TAM for agentic AI (~$24.5B–$139B by the early 2030s) measures the whole market — "
             "mostly US, mostly SaaS, mostly the model and agent-framework layers ABOM deliberately does not sell. "
             "This model sizes only what ABOM can actually sell: the agent-accountability / governance tooling layer "
             "(generate, verify, and notarize a signed bill of materials for every agent) for European regulated buyers. "
             "That layer rides the SBOM-mandate precedent — EO 14028 and the EU Cyber Resilience Act made software bills "
             "of materials mandatory, and EU AI Act Art. 12, DORA, and NIST AI RMF point the same way for agents. "
             "Blue = input you can change; black = formula; green = pulled from another sheet.")
ws["A25"].font = f_note
ws["A25"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A25:I27")

# =====================================================================
# SHEET 2: ASSUMPTIONS
# =====================================================================
wa = wb.create_sheet("Assumptions")
wa.sheet_view.showGridLines = False
wa["A1"] = "Assumptions"; wa["A1"].font = f_title
wa["A2"] = "Blue cells are inputs — change them to flex the model. Currency in € thousands (€000)."; wa["A2"].font = f_sub

# Section A: segment universe & SAM
wa["A4"] = "A. Serviceable market by segment"; wa["A4"].font = f_sect
for c in range(1,7):
    wa.cell(row=4, column=c).fill = fill_sect
hdrs = ["Segment", "EU institution\nuniverse (#)", "Serviceable\nshare (%)", "Serviceable\nlogos (#)", "Mature ACV\n(€000)", "Segment SAM\n(€000)"]
style_row_headers(wa, 5, [1,2,3,4,5,6])
for i, h in enumerate(hdrs):
    wa.cell(row=5, column=1+i, value=h)

# rows: (name, universe, serv%, ACV)  -- FS beachhead then expansion
fs = [
    ("Tier-1 banks (G-SIB / ECB significant)", 120, 0.70, 1200),
    ("Tier-2 banks (mid-size / regional)", 400, 0.40, 400),
    ("Large insurers", 150, 0.50, 500),
    ("Asset mgmt / payments / market infra", 250, 0.35, 350),
]
exp = [
    ("Defence & public sector", 300, 0.40, 650),
    ("Healthcare (large systems / pharma)", 250, 0.30, 450),
    ("Telecom", 80, 0.50, 550),
]

def seg_row(r, name, universe, serv, acv, zebra):
    wa.cell(row=r, column=1, value=name).font = f_lbl
    b = wa.cell(row=r, column=2, value=universe); b.font = f_inp; b.number_format = CNT
    c = wa.cell(row=r, column=3, value=serv); c.font = f_inp; c.number_format = PCT
    d = wa.cell(row=r, column=4, value=f"=B{r}*C{r}"); d.font = f_calc; d.number_format = CNT
    e = wa.cell(row=r, column=5, value=acv); e.font = f_inp; e.number_format = CUR
    fcell = wa.cell(row=r, column=6, value=f"=D{r}*E{r}"); fcell.font = f_calc; fcell.number_format = CUR
    for col in range(1,7):
        cell = wa.cell(row=r, column=col); cell.border = border_all
        if zebra: cell.fill = fill_zeb

r = 6
for i,(n,u,s,a) in enumerate(fs):
    seg_row(r, n, u, s, a, i % 2 == 1); r += 1
# subtotal FS
wa.cell(row=r, column=1, value="Subtotal — Financial services (beachhead)").font = f_bold
wa.cell(row=r, column=4, value=f"=SUM(D6:D9)").font = f_bold; wa.cell(row=r, column=4).number_format = CNT
wa.cell(row=r, column=6, value=f"=SUM(F6:F9)").font = f_bold; wa.cell(row=r, column=6).number_format = CUR
for col in range(1,7): wa.cell(row=r, column=col).border = border_all
fs_sub = r; r += 1
exp_start = r
for i,(n,u,s,a) in enumerate(exp):
    seg_row(r, n, u, s, a, i % 2 == 1); r += 1
exp_end = r-1
wa.cell(row=r, column=1, value="Subtotal — Expansion verticals").font = f_bold
wa.cell(row=r, column=4, value=f"=SUM(D{exp_start}:D{exp_end})").font = f_bold; wa.cell(row=r, column=4).number_format = CNT
wa.cell(row=r, column=6, value=f"=SUM(F{exp_start}:F{exp_end})").font = f_bold; wa.cell(row=r, column=6).number_format = CUR
for col in range(1,7): wa.cell(row=r, column=col).border = border_all
exp_sub = r; r += 1
# TOTAL SAM
wa.cell(row=r, column=1, value="TOTAL serviceable market (SAM)").font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
tc = wa.cell(row=r, column=4, value=f"=D{fs_sub}+D{exp_sub}"); tc.font = Font(name=ARIAL, bold=True, color="FFFFFF"); tc.number_format = CNT
tf = wa.cell(row=r, column=6, value=f"=F{fs_sub}+F{exp_sub}"); tf.font = Font(name=ARIAL, bold=True, color="FFFFFF"); tf.number_format = CUR
for col in range(1,7):
    cell = wa.cell(row=r, column=col); cell.fill = fill_card; cell.border = border_all
SAM_LOGOS_ROW = r  # D
SAM_EUR_ROW = r    # F
sam_row = r

# segment ACV rows map (for revenue build): FS rows 6-9, exp rows exp_start..exp_end
acv_rows = [6,7,8,9, exp_start, exp_start+1, exp_start+2]

# Section B: global drivers
br = sam_row + 2
wa.cell(row=br, column=1, value="B. Global drivers").font = f_sect
for c in range(1,7): wa.cell(row=br, column=c).fill = fill_sect
wa.cell(row=br+1, column=1, value="ACV expansion / escalation (per year)").font = f_lbl
g_cell = wa.cell(row=br+1, column=2, value=0.08); g_cell.font = f_inp; g_cell.number_format = PCT
wa.cell(row=br+2, column=1, value="Services attach (% of platform ARR)").font = f_lbl
sv_cell = wa.cell(row=br+2, column=2, value=0.20); sv_cell.font = f_inp; sv_cell.number_format = PCT
G_ROW = br+1
SV_ROW = br+2

# Section C: adoption — cumulative active logos by year
cr = br + 4
wa.cell(row=cr, column=1, value="C. Adoption — cumulative active customers (logos) by year").font = f_sect
for c in range(1,7): wa.cell(row=cr, column=c).fill = fill_sect
style_row_headers(wa, cr+1, [1,2,3,4,5,6])
wa.cell(row=cr+1, column=1, value="Segment")
for i in range(5):
    wa.cell(row=cr+1, column=2+i, value=f"Year {i+1}")
# cumulative logos inputs
logos = [
    ("Tier-1 banks", [1,2,4,8,14]),
    ("Tier-2 banks", [0,1,4,10,22]),
    ("Large insurers", [0,1,3,6,12]),
    ("Asset mgmt / payments / market infra", [0,0,1,3,7]),
    ("Defence & public sector", [0,0,0,1,4]),
    ("Healthcare", [0,0,0,1,3]),
    ("Telecom", [0,0,0,0,2]),
]
logo_start = cr+2
rr = logo_start
for i,(n,vals) in enumerate(logos):
    wa.cell(row=rr, column=1, value=n).font = f_lbl
    for j,v in enumerate(vals):
        cell = wa.cell(row=rr, column=2+j, value=v); cell.font = f_inp; cell.number_format = CNT
        cell.border = border_all
    wa.cell(row=rr, column=1).border = border_all
    if i % 2 == 1:
        for col in range(1,7): wa.cell(row=rr, column=col).fill = fill_zeb
    rr += 1
logo_end = rr-1
# total logos
wa.cell(row=rr, column=1, value="Total cumulative customers").font = f_bold
for j in range(5):
    col = 2+j
    L = get_column_letter(col)
    cell = wa.cell(row=rr, column=col, value=f"=SUM({L}{logo_start}:{L}{logo_end})")
    cell.font = f_bold; cell.number_format = CNT; cell.border = border_all
wa.cell(row=rr, column=1).border = border_all
TOTAL_LOGO_ROW = rr

# widths
wa.column_dimensions["A"].width = 40
for col in "BCDEF":
    wa.column_dimensions[col].width = 14
wa.row_dimensions[5].height = 30
wa.row_dimensions[cr+1].height = 18

# stash references for other sheets
refs = dict(sam_row=sam_row, acv_rows=acv_rows, g_row=G_ROW, sv_row=SV_ROW,
            logo_start=logo_start, total_logo_row=TOTAL_LOGO_ROW)

# =====================================================================
# SHEET 3: MARKET SIZING
# =====================================================================
wm = wb.create_sheet("Market Sizing")
wm.sheet_view.showGridLines = False
wm["A1"] = "Market Sizing — ABOM TAM / SAM / SOM"; wm["A1"].font = f_title
wm["A2"] = "Top-down TAM funnel triangulated against bottom-up SAM. Figures in € millions (€M)."; wm["A2"].font = f_sub

wm["A4"] = "Top-down TAM funnel (€M)"; wm["A4"].font = f_sect
for c in range(1,4): wm.cell(row=4, column=c).fill = fill_sect

# rows
def ms_row(r, label, value=None, formula=None, fmt=CURM, font=None, note=None):
    wm.cell(row=r, column=1, value=label).font = font or f_lbl
    if value is not None:
        cell = wm.cell(row=r, column=2, value=value); cell.font = f_inp
    else:
        cell = wm.cell(row=r, column=2, value=formula); cell.font = f_calc
    cell.number_format = fmt
    if note:
        wm.cell(row=r, column=3, value=note).font = f_note

ms_row(5, "Global agentic-AI market, 2030E (directional midpoint)", value=46000, fmt=CURM,
       note="~$50B; analyst range $24.5B (GVR) – $139B (Fortune). Directional.")
ms_row(6, "× Agent-accountability / governance tooling-layer share", value=0.15, fmt=PCT,
       note="Accountability/governance tooling layer as share of total agentic spend (assumption).")
ms_row(7, "= Global agent-accountability TAM", formula="=B5*B6")
ms_row(8, "× Europe share", value=0.22, fmt=PCT, note="EU share of global enterprise AI spend (assumption).")
ms_row(9, "= European agent-accountability TAM", formula="=B7*B8")
ms_row(10, "= EU regulated / mandate-driven share", value=0.45, fmt=PCT,
       note="Share of EU accountability layer in regulated, mandate-exposed segments (assumption).")
# fix: row10 should be a multiplier line then product; restructure
wm.cell(row=10, column=1).value = "× EU regulated / mandate-driven share"
wm.cell(row=11, column=1, value="= EU regulated agent-accountability TAM (top-down)").font = f_bold
tc = wm.cell(row=11, column=2, value="=B9*B10"); tc.font = f_bold; tc.number_format = CURM
# Re-point summary card (TAM) to B11 — adjust: cards used B10. Update card formula now.
ws["A4"].value = "TAM — EU regulated agent-accountability layer (top-down)"
ws["A5"].value = "='Market Sizing'!B11"

# Bottom-up SAM & SOM
wm["A13"] = "Bottom-up cross-check (€M)"; wm["A13"].font = f_sect
for c in range(1,4): wm.cell(row=13, column=c).fill = fill_sect
sam_r = refs["sam_row"]
wm.cell(row=14, column=1, value="SAM — bottom-up (serviceable, annual)").font = f_lbl
sc = wm.cell(row=14, column=2, value=f"=Assumptions!F{sam_r}/1000"); sc.font = f_link; sc.number_format = CURM
wm.cell(row=14, column=3, value="From Assumptions §A (serviceable logos × mature ACV).").font = f_note
wm.cell(row=15, column=1, value="SOM — Year-5 achievable ARR (base case)").font = f_lbl
oc = wm.cell(row=15, column=2, value="='Revenue Build'!G12/1000"); oc.font = f_link; oc.number_format = CURM
wm.cell(row=15, column=3, value="From Revenue Build (Year-5 ABOM platform ARR).").font = f_note
wm.cell(row=16, column=1, value="SOM as % of SAM").font = f_lbl
pc = wm.cell(row=16, column=2, value="=B15/B14"); pc.font = f_calc; pc.number_format = PCT
wm.cell(row=17, column=1, value="SAM as % of EU regulated TAM (top-down)").font = f_lbl
qc = wm.cell(row=17, column=2, value="=B14/B11"); qc.font = f_calc; qc.number_format = PCT

# fix summary SAM/SOM card formulas to point at B14/B15
ws["D5"].value = "='Market Sizing'!B14"
ws["G5"].value = "='Market Sizing'!B15"

wm.column_dimensions["A"].width = 48
wm.column_dimensions["B"].width = 14
wm.column_dimensions["C"].width = 50

# =====================================================================
# SHEET 4: REVENUE BUILD
# =====================================================================
wr = wb.create_sheet("Revenue Build")
wr.sheet_view.showGridLines = False
wr["A1"] = "Revenue Build — ABOM 5-Year Bottom-Up"; wr["A1"].font = f_title
wr["A2"] = ("ABOM platform ARR (€000) = cumulative customers × mature ACV × (1+expansion)^(year-1); "
            "ACV = abom-verify + Notary subscription per logo. All drivers live on the Assumptions sheet."); wr["A2"].font = f_sub

style_row_headers(wr, 4, [1,2,3,4,5,6])
wr.cell(row=4, column=1, value="ABOM platform ARR by segment (€000)")
for i in range(5):
    wr.cell(row=4, column=2+i, value=f"Year {i+1}")

seg_names = [n for n,_ in logos]
acv_rows = refs["acv_rows"]
logo_start = refs["logo_start"]
g_row = refs["g_row"]
sv_row = refs["sv_row"]
total_logo_row = refs["total_logo_row"]

rb_start = 5
for i, name in enumerate(seg_names):
    rrow = rb_start + i
    wr.cell(row=rrow, column=1, value=name).font = f_lbl
    logo_r = logo_start + i
    acv_r = acv_rows[i]
    for j in range(5):
        col = 2+j
        L = get_column_letter(col)
        # cumulative logos for this year × ACV × (1+g)^(year-1)
        formula = (f"=Assumptions!{L}{logo_r}*Assumptions!$E${acv_r}"
                   f"*(1+Assumptions!$B${g_row})^{j}")
        cell = wr.cell(row=rrow, column=col, value=formula)
        cell.font = f_calc; cell.number_format = CUR; cell.border = border_all
    wr.cell(row=rrow, column=1).border = border_all
    if i % 2 == 1:
        for col in range(1,7): wr.cell(row=rrow, column=col).fill = fill_zeb
rb_end = rb_start + len(seg_names) - 1

# total platform ARR (row 12)
tr = rb_end + 1
wr.cell(row=tr, column=1, value="Total ABOM platform ARR").font = f_bold
for j in range(5):
    col = 2+j; L = get_column_letter(col)
    cell = wr.cell(row=tr, column=col, value=f"=SUM({L}{rb_start}:{L}{rb_end})")
    cell.font = f_bold; cell.number_format = CUR; cell.border = border_all
wr.cell(row=tr, column=1).border = border_all
TOTAL_ARR_ROW = tr  # =12 expected

# services (row 13)
svr = tr + 1
wr.cell(row=svr, column=1, value="Services revenue").font = f_lbl
for j in range(5):
    col = 2+j; L = get_column_letter(col)
    cell = wr.cell(row=svr, column=col, value=f"={L}{tr}*Assumptions!$B${sv_row}")
    cell.font = f_calc; cell.number_format = CUR; cell.border = border_all
wr.cell(row=svr, column=1).border = border_all

# total revenue (row 14)
tvr = svr + 1
wr.cell(row=tvr, column=1, value="Total revenue").font = f_bold
for j in range(5):
    col = 2+j; L = get_column_letter(col)
    cell = wr.cell(row=tvr, column=col, value=f"={L}{tr}+{L}{svr}")
    cell.font = f_bold; cell.number_format = CUR; cell.border = border_all
    cell.fill = fill_zeb
wr.cell(row=tvr, column=1).border = border_all; wr.cell(row=tvr, column=1).fill = fill_zeb

# spacer + logo metrics
mr = tvr + 2
wr.cell(row=mr, column=1, value="Cumulative customers (logos)").font = f_lbl
for j in range(5):
    col = 2+j; L = get_column_letter(col)
    cell = wr.cell(row=mr, column=col, value=f"=Assumptions!{L}{total_logo_row}")
    cell.font = f_link; cell.number_format = CNT
LOGO_METRIC_ROW = mr  # =16 expected
pr = mr + 1
wr.cell(row=pr, column=1, value="Logo penetration of serviceable market").font = f_lbl
for j in range(5):
    col = 2+j; L = get_column_letter(col)
    cell = wr.cell(row=pr, column=col, value=f"={L}{mr}/Assumptions!$D${refs['sam_row']}")
    cell.font = f_calc; cell.number_format = PCT
PEN_ROW = pr  # =17 expected

# ARR in €M helper row
amr = pr + 2
wr.cell(row=amr, column=1, value="Memo: Total ABOM platform ARR (€M)").font = f_note
for j in range(5):
    col = 2+j; L = get_column_letter(col)
    cell = wr.cell(row=amr, column=col, value=f"={L}{tr}/1000")
    cell.font = f_note; cell.number_format = CURM

wr.column_dimensions["A"].width = 40
for col in "BCDEFG":
    wr.column_dimensions[col].width = 13
wr.row_dimensions[4].height = 18

# verify expected rows for cross-sheet refs
assert TOTAL_ARR_ROW == 12, TOTAL_ARR_ROW
assert LOGO_METRIC_ROW == 16, LOGO_METRIC_ROW
assert PEN_ROW == 17, PEN_ROW

try:
    wb.calculation.fullCalcOnLoad = True   # force recalc when opened
except Exception:
    pass
wb.save("/Users/josephassiga/Desktop/aegis/ABOM_Market_Model.xlsx")
print("saved. TOTAL_ARR_ROW", TOTAL_ARR_ROW, "SAM_ROW", refs["sam_row"], "TOTAL_LOGO_ROW", total_logo_row)
